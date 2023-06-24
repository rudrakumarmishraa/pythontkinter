#Importing Packages Used
from tkinter import *
from tkinter import ttk
from tkinter import messagebox ,simpledialog
from PIL import Image, ImageTk
import openpyxl as xl
from os import getlogin

user = getlogin()

class Student:
    def __init__(self, root):
        #Defining Window Attributes
        self.root = root
        self.root.geometry("1350x600+0+50")
        self.root.minsize(1350, 600)
        self.root.maxsize(1350, 600)
        self.root["bg"]="white"
        self.root.title("Student Management System      --Rudra Kumar Mishra")

        # Heading of the Window
        lblHeading = Label(self.root, text = 'Student Management System', font = ('Algerian', 52, 'bold'), bg = 'white', fg = 'red')
        lblHeading.place(x = 0, y = 0, height = 70, width = 1300)

        # Creating Main Frame to Enter Widgets
        mainframe = Frame(self.root, border = 10, relief = SUNKEN)
        mainframe.place(x = 0, y = 80, width = 1350, height = 515)


        # Left Frame for Student Details
        lefframe = LabelFrame(mainframe, border = 10, relief = RIDGE, text = 'Add Student Information', font = ('Algerian', 12, 'bold'), bg = 'white')
        lefframe.place(x = 5, y = 5, width = 680, height = 480)



        # Course Information Frame with four Comboboxes
        Courseframe = LabelFrame(lefframe, border = 5, relief = RIDGE, text = 'Current Course Information', font = ('Copper Black', 12), bg = 'white')
        Courseframe.place(x = 0, y = 0, width = 660, height = 150)

        #   Department Label and Combobox
        lblDepartement = Label(Courseframe, width = 10, font = ('Copper Black', 12, 'bold'), text = 'Department', bg = 'white')
        lblDepartement.grid(row = 0, column = 0)
        self.Department = ttk.Combobox(Courseframe, font = ('Copper Black', 12), state = 'readonly')
        self.Department["values"]=('Select Department', 'Education', 'Technology')
        self.Department.current(0)
        self.Department.grid(row=0, column=2, padx = 2, pady = 10, sticky = W)

        #   Course Label and Combobox
        lblCourse = Label(Courseframe, width = 10, font = ('Copper Black', 12, 'bold'), text = 'Course', bg = 'white')
        lblCourse.grid(row = 0, column = 3)
        self.Course = ttk.Combobox(Courseframe, font = ('Copper Black', 12), state = 'readonly')
        self.Course["values"]=('Select Course', 'MCA','MBA','MCom', 'MEd', 'BCA', 'BBA', 'BCom', 'BEd')
        self.Course.current(0)
        self.Course.grid(row=0, column=4, padx = 2, pady = 10, sticky = W)

        #   Year Label and Combobox
        lblYear = Label(Courseframe, width = 10, font = ('Copper Black', 12, 'bold'), text = 'Year', bg = 'white')
        lblYear.grid(row = 1, column = 0)
        self.Year = ttk.Combobox(Courseframe, font = ('Copper Black', 12), state = 'readonly')
        self.Year["values"]=('Select Year', 'First Year', 'Second Year', 'Third Year')
        self.Year.current(0)
        self.Year.grid(row=1, column=2, padx = 2, pady = 10, sticky = W)

        #   Semester Label and Combobox
        lblSemester = Label(Courseframe, width = 10, font = ('Copper Black', 12, 'bold'), text = 'Semester', bg = 'white')
        lblSemester.grid(row = 1, column = 3)
        self.Semester = ttk.Combobox(Courseframe, font = ('Copper Black', 12), state = 'readonly')
        self.Semester["values"]=('Select Semester', 'First Semester', 'Second Semester', 'Third Semester', 'Fourth Semester', 'Fifth Semester', 'Sixth Semester')
        self.Semester.current(0)
        self.Semester.grid(row=1, column=4, padx = 2, pady = 10, sticky = W)



        # Class and Student Information
        StuentInfoFrame = LabelFrame(lefframe, border = 5, relief = RIDGE, text = 'Class Student Information', font = ('Copper Black', 12), bg = 'white')
        StuentInfoFrame.place(x = 0, y = 150, width = 660, height = 200)

        #Student ID
        self.StudentIDvar = StringVar()
        lblStudentID = Label(StuentInfoFrame, text = 'Student ID', font = ('Copper Black', 12, 'bold'), bg = 'white')
        lblStudentID.grid(row = 0, column = 0, padx = 10, pady = 10)
        self.StudentID = ttk.Entry(StuentInfoFrame, font = ('Copper Black', 12), textvariable = self.StudentIDvar)
        self.StudentID.grid(row = 0, column = 1, padx = 2, pady = 10)

        # Student Gender
        lblGender = Label(StuentInfoFrame, text = 'Gender', font = ('Copper Black', 12, 'bold'), bg = 'white')
        lblGender.grid(row = 1, column = 0, padx = 10, pady = 10)
        self.Gender = ttk.Combobox(StuentInfoFrame, font = ('Copper Black', 11), state = 'readonly')
        self.Gender["values"]=('Select Gender', 'MALE', 'FEMALE')
        self.Gender.current(0)
        self.Gender.grid(row=1, column=1, padx = 2, pady = 10, sticky = W)

        # Student Address
        self.StudentAddressvar = StringVar()
        lblAddress = Label(StuentInfoFrame, text = 'Address', font = ('Copper Black', 12, 'bold'), bg = 'white')
        lblAddress.grid(row = 2, column = 0, padx = 10, pady = 10)
        self.Address = ttk.Entry(StuentInfoFrame, font = ('Copper Black', 12), textvariable = self.StudentAddressvar)
        self.Address.grid(row = 2, column = 1, padx = 2, pady = 10)

        # Student Name
        self.StudentNamevar = StringVar()
        lblName = Label(StuentInfoFrame, text = 'Student Name', font = ('Copper Black', 12, 'bold'), bg = 'white')
        lblName.grid(row = 0, column =3, padx = 10, pady = 10)
        self.Name = ttk.Entry(StuentInfoFrame, font = ('Copper Black', 12), textvariable = self.StudentNamevar)
        self.Name.grid(row = 0, column = 4, padx = 2, pady = 10)

        # Student Date of Birth
        self.Date_of_Birthvar = StringVar()
        lblDate_of_Birth = Label(StuentInfoFrame, text = 'Date of Birth', font = ('Copper Black', 12, 'bold'), bg = 'white')
        lblDate_of_Birth.grid(row = 1, column = 3, padx = 10, pady = 10)
        self.Date_of_Birth = ttk.Entry(StuentInfoFrame, font = ('Copper Black', 12), textvariable = self.Date_of_Birthvar)
        self.Date_of_Birth.grid(row = 1, column = 4, padx = 2, pady = 10)

        # Student Phone_Number
        self.PhoneNumbervar = StringVar()
        lblPhone_Number = Label(StuentInfoFrame, text = 'Phone Number', font = ('Copper Black', 12, 'bold'), bg = 'white')
        lblPhone_Number.grid(row = 2, column = 3, padx = 10, pady = 10)
        self.Phone_Number = ttk.Entry(StuentInfoFrame, font = ('Copper Black', 12), textvariable = self.PhoneNumbervar)
        self.Phone_Number.grid(row = 2, column = 4, padx = 2, pady = 10)

        # Radio Buttons
        self.Radio = StringVar()
        self.RadioButton1 = ttk.Radiobutton(StuentInfoFrame, text = 'Take Photo Sample', value = 'Yes', variable = self.Radio)
        self.RadioButton1.grid(row = 3, column = 0)
        self.RadioButton2 = ttk.Radiobutton(StuentInfoFrame, text = 'Do not take Photo Sample', value = 'No', variable = self.Radio)
        self.RadioButton2.grid(row = 3, column = 1)



        # Button for Functionality of Program
        ButtonFrame = Frame(lefframe, bd = 5, relief = RIDGE, bg = 'white')
        ButtonFrame.place(x = 3, y = 355, height = 90, width = 660)

        #Add Data Button
        save_btn = Button(ButtonFrame, text = 'Save', command = self.insertdata, font = ('Copper Black', 13, 'bold'), bg = 'blue', fg = 'white', width = 15, height = 3)
        save_btn.grid(row = 0, column = 0, padx = 1, pady = 5)
        #Reset Fields Button
        Reset_btn = Button(ButtonFrame, text = 'Reset', command = self.reset, font = ('Copper Black', 13, 'bold'), bg = 'blue', fg = 'white', width = 15, height = 3)
        Reset_btn.grid(row = 0, column = 1, padx = 1, pady = 5)
        #Train Data Button
        Show_All_btn = Button(ButtonFrame, text = 'Show Table', font = ('Copper Black', 13, 'bold'), bg = 'blue', fg = 'white', width = 15, height = 3, command = self.showall)
        Show_All_btn.grid(row = 0, column = 2, padx = 1, pady = 5)
        #Face Detector Button
        Face_Detector_btn = Button(ButtonFrame, text = 'Quit', command = self.root.destroy, font = ('Copper Black', 13, 'bold'), bg = 'blue', fg = 'white', width = 15, height = 3)
        Face_Detector_btn.grid(row = 0, column =3, padx = 1, pady = 5)

        # Left Frame GUI Ends Here



        # Right Frame for Student Details
        rightframe = LabelFrame(mainframe, border = 10, relief = RIDGE, text = 'Student Details', font = ('Algerian', 12, 'bold'), bg = 'white')
        rightframe.place(x = 690, y = 5, width = 635, height = 480)


        #Serch Frame For Student Info
        Seach_Frame = LabelFrame(rightframe, border = 10, relief = RIDGE, text = 'Seach Student Record', font = ('Algerian', 10), bg = 'white')
        Seach_Frame.place(x = 5, y = 5, width = 610, height = 80)

        lblSeachby = Label(Seach_Frame, text = 'Seach By', font = ('Lucida Handwritting', 16), bg = 'red', fg = 'white')
        lblSeachby.grid(row = 0, column = 0, pady = 10, padx = 10)

        self.Seachby = ttk.Combobox(Seach_Frame, font = ('Copper Black', 12), state = 'readonly')
        self.Seachby["values"]=('Select By', 'Course', 'Student ID')
        self.Seachby.current(0)
        self.Seachby.grid(row=0, column=1, sticky = W)

        self.Seach_Data = ttk.Entry(Seach_Frame, font = ('Copper Black', 12))
        self.Seach_Data.grid(row = 0, column = 2, padx = 5)

        # Seach Button
        Search_btn = Button(Seach_Frame, text = 'Seach', font = ('Copper Black', 13, 'bold'), bg = 'blue', fg = 'white', command = self.search)
        Search_btn.grid(row = 0, column = 3, padx = 5)

        # Student Record
        Student_Table_Record = LabelFrame(rightframe, border = 10, relief = RIDGE, text = 'Student Table', font = ('Algerian', 10), bg = 'white')
        Student_Table_Record.place(x = 5, y = 90, width = 610, height = 355)


        # Student Record Table
        col = ('Department', 'Course', 'Year', 'Semester', 'Student ID', 'Student Name', 'Gender',   'Address', 'Phone Number', 'Date of Birth', 'Photo Sample')
        scrool_y = Scrollbar(Student_Table_Record, orient=VERTICAL)
        scrool_x = Scrollbar(Student_Table_Record, orient=HORIZONTAL)
        self.student_table = ttk.Treeview(Student_Table_Record, height=200, show='headings', columns=col, yscrollcommand = scrool_y.set, xscrollcommand = scrool_x.set)

        self.student_table.column('Department',width = 100, anchor=CENTER, minwidth = 100)
        self.student_table.column('Course', width = 100, anchor=CENTER, minwidth = 100)
        self.student_table.column('Year', width = 100, anchor=CENTER, minwidth = 100)
        self.student_table.column('Semester', width = 100, anchor=CENTER, minwidth = 100)
        self.student_table.column('Student ID', width = 100, anchor=CENTER, minwidth = 100)
        self.student_table.column('Student Name', width = 100, anchor=CENTER, minwidth = 100)
        self.student_table.column('Gender', width = 100, anchor=CENTER, minwidth = 100)
        self.student_table.column('Address', width = 100, anchor=CENTER, minwidth = 100)
        self.student_table.column('Phone Number', width = 100, anchor=CENTER, minwidth = 100)
        self.student_table.column('Date of Birth', width = 100, anchor=CENTER, minwidth = 100)
        self.student_table.column('Photo Sample', width = 100, anchor=CENTER, minwidth = 100)

        self.student_table.heading('Department', text = 'Department')
        self.student_table.heading('Course', text = 'Course')
        self.student_table.heading('Year', text = 'Year')
        self.student_table.heading('Semester', text = 'Semester')
        self.student_table.heading('Student ID', text = 'Student ID')
        self.student_table.heading('Student Name', text = 'Student Name')
        self.student_table.heading('Gender', text = 'Gender')
        self.student_table.heading('Address', text = 'Address')
        self.student_table.heading('Phone Number', text = 'Phone Number')
        self.student_table.heading('Date of Birth', text = 'Date of Birth')
        self.student_table.heading('Photo Sample', text = 'Photo Sample')

        scrool_y.pack(side = RIGHT, fill=Y)
        scrool_x.pack(side = BOTTOM, fill=X)
        self.student_table.pack(side=TOP, fill=Y)
        scrool_y.config(command=self.student_table.yview)
        scrool_x.config(command=self.student_table.xview)
    
    def insertdata(self):
        if self.Radio.get() == '' or self.Department.get() == 'Select Department' or self.Course.get() == 'Select Course' or self.Year.get() == 'Select Year' or self.Semester.get() == 'Select Semester' or self.StudentIDvar.get() == '' or self.StudentNamevar.get() == '' or self.Gender.get() == 'Select Gender' or self.Date_of_Birthvar.get() == '' or self.StudentAddressvar.get() == '' or self.PhoneNumbervar.get() == '':
            messagebox.showerror('All Fields Required', 'Please Enter All Fields to Save Data.\t\n')

        else:
            try:
                wb = xl.load_workbook("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")
                sheetMCA = wb["MCA"]
                sheetMBA = wb["MBA"]
                sheetMCom = wb["MCom"]
                sheetMEd = wb["MEd"]
                sheetBCA = wb["BCA"]
                sheetBBA = wb["BBA"]
                sheetBCom = wb["BCom"]
                sheetBEd = wb["BEd"]
                sheettup = ('MCA', 'MBA', 'MCom', 'MEd', 'BCA', 'BBA', 'BCom', 'BEd')
                tempsheettup = (sheetMCA, sheetMBA, sheetMCom, sheetMEd, sheetBCA, sheetBBA, sheetBCom, sheetBEd)
                temptup = ('Department', 'Course', 'Year', 'Semester', 'Student ID', 'Student Name', 'Gender', 'Date of Birth', 'Address', 'Phone Number', 'Photo Sample')
                datatup = (self.Department.get(), self.Course.get(), self.Year.get(), self.Semester.get(), self.StudentIDvar.get(), self.StudentNamevar.get(), self.Gender.get(), self.Date_of_Birthvar.get(), self.StudentAddressvar.get(), self.PhoneNumbervar.get(), self.Radio.get())

                a = 1
                x = 2
                self.student_table.delete(*self.student_table.get_children())
                for i in range(8):
                    while tempsheettup[i].cell(x, 1).value != None:
                        if tempsheettup[i].cell(x, 5).value == datatup[4]:
                            messagebox.showwarning('Warning', 'Data not added because Student ID already Exists.\t\n')
                            return None
                        x+=1
                        a+=1
                    x = 2

                x = 1
                for i in range(1, 9):
                    if sheettup[i-1] == datatup[1]:
                        while tempsheettup[i-1].cell(x, 1).value != None:
                            x+=1
                        for j in range(1, 12):
                                tempsheettup[i-1].cell(x, j).value = datatup[j-1]
                wb.save("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")
                messagebox.showinfo('Success', 'Added data Successfully!')
                self.showall()

            except FileNotFoundError:
                # If file not found then making a workbook
                wb = xl.Workbook("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")
                sheetMCA = wb.create_sheet("MCA")
                sheetMBA = wb.create_sheet("MBA")
                sheetMCom = wb.create_sheet("MCom")
                sheetMEd = wb.create_sheet("MEd")
                sheetBCA = wb.create_sheet("BCA")
                sheetBBA = wb.create_sheet("BBA")
                sheetBCom = wb.create_sheet("BCom")
                sheetBEd = wb.create_sheet("BEd")
                wb.save("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")
                #WorkBook Saved and then opened to Write
                wb = xl.load_workbook("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")
                sheetMCA = wb["MCA"]
                sheetMBA = wb["MBA"]
                sheetMCom = wb["MCom"]
                sheetMEd = wb["MEd"]
                sheetBCA = wb["BCA"]
                sheetBBA = wb["BBA"]
                sheetBCom = wb["BCom"]
                sheetBEd = wb["BEd"]
                sheettup = ('MCA', 'MBA', 'MCom', 'MEd', 'BCA', 'BBA', 'BCom', 'BEd')
                tempsheettup = (sheetMCA, sheetMBA, sheetMCom, sheetMEd, sheetBCA, sheetBBA, sheetBCom, sheetBEd)
                temptup = ('Department', 'Course', 'Year', 'Semester', 'Student ID', 'Student Name', 'Gender', 'Date of Birth', 'Address', 'Phone Number', 'Photo Sample')
                datatup = (self.Department.get(), self.Course.get(), self.Year.get(), self.Semester.get(), self.StudentIDvar.get(), self.StudentNamevar.get(), self.Gender.get(), self.Date_of_Birthvar.get(), self.StudentAddressvar.get(), self.PhoneNumbervar.get(), self.Radio.get())
                for i in range(1, 9):
                    for j in range(1, 12):
                        tempsheettup[i-1].cell(1, j).value = temptup[j-1]
                        if sheettup[i-1] == datatup[1]:
                            tempsheettup[i-1].cell(2, j).value = datatup[j-1]
                wb.save("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")
                messagebox.showinfo('Great Success', 'Workbook created and data saved successfully.\t\t\n')
                self.showall()
                #WorkBook Saved after Creating First Entry 
    
    def reset(self):
        self.Department.set('Select Department')
        self.Course.set('Select Course')
        self.Year.set('Select Year')
        self.Semester.set('Select Semester')
        self.Gender.set('Select Gender')
        self.StudentNamevar.set('')
        self.PhoneNumbervar.set('')
        self.StudentAddressvar.set('')
        self.StudentIDvar.set('')
        self.Date_of_Birthvar.set('')

    def showall(self):
        try:
            wb = xl.load_workbook("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")
            sheetMCA = wb["MCA"]
            sheetMBA = wb["MBA"]
            sheetMCom = wb["MCom"]
            sheetMEd = wb["MEd"]
            sheetBCA = wb["BCA"]
            sheetBBA = wb["BBA"]
            sheetBCom = wb["BCom"]
            sheetBEd = wb["BEd"]
            tempsheettup = (sheetMCA, sheetMBA, sheetMCom, sheetMEd, sheetBCA, sheetBBA, sheetBCom, sheetBEd)

            a = 1
            x = 2
            self.student_table.delete(*self.student_table.get_children())
            for i in range(8):
                while tempsheettup[i].cell(x, 1).value != None:
                    val = (tempsheettup[i].cell(x, 1).value, tempsheettup[i].cell(x, 2).value, tempsheettup[i].cell(x, 3).value, tempsheettup[i].cell(x, 4).value, tempsheettup[i].cell(x, 5).value, tempsheettup[i].cell(x, 6).value, tempsheettup[i].cell(x, 7).value, tempsheettup[i].cell(x, 8).value, tempsheettup[i].cell(x, 9).value, tempsheettup[i].cell(x, 10).value, tempsheettup[i].cell(x, 11).value)
                    self.student_table.insert('', a, values = val)
                    x+=1
                    a+=1
                x = 2
                    
            wb.save("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")

        except FileNotFoundError:
            messagebox.showerror('No Data Found', 'No Workbook is present Add Some Data.\t\t\n') 

    def search(self):
        if self.Seachby.get() == 'Course':
            try:
                wb = xl.load_workbook("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")

                sheet = wb[self.Seach_Data.get()]

                a = 1
                x = 2
                self.student_table.delete(*self.student_table.get_children())
                while sheet.cell(x, 1).value != None:
                    val = (sheet.cell(x, 1).value, sheet.cell(x, 2).value, sheet.cell(x, 3).value, sheet.cell(x, 4).value, sheet.cell(x, 5).value, sheet.cell(x, 6).value, sheet.cell(x, 7).value, sheet.cell(x, 8).value, sheet.cell(x, 9).value, sheet.cell(x, 10).value, sheet.cell(x, 11).value)
                    self.student_table.insert('', a, values = val)
                    x+=1
                    a+=1
                        
                wb.save("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")

            except FileNotFoundError:
                messagebox.showerror('No Data Found', 'No Workbook is present Add Some Data.\t\t\n') 
            except KeyError:
                messagebox.showerror('No Data Found', 'No Such Course is present.\t\t\n') 

        elif self.Seachby.get() == 'Student ID':
            try:
                wb = xl.load_workbook("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")
                sheetMCA = wb["MCA"]
                sheetMBA = wb["MBA"]
                sheetMCom = wb["MCom"]
                sheetMEd = wb["MEd"]
                sheetBCA = wb["BCA"]
                sheetBBA = wb["BBA"]
                sheetBCom = wb["BCom"]
                sheetBEd = wb["BEd"]
                tempsheettup = (sheetMCA, sheetMBA, sheetMCom, sheetMEd, sheetBCA, sheetBBA, sheetBCom, sheetBEd)

                a = 1
                x = 2
                self.student_table.delete(*self.student_table.get_children())
                for i in range(8):
                    while tempsheettup[i].cell(x, 1).value != None:
                        if tempsheettup[i].cell(x, 5).value == self.Seach_Data.get():
                            val = (tempsheettup[i].cell(x, 1).value, tempsheettup[i].cell(x, 2).value, tempsheettup[i].cell(x, 3).value, tempsheettup[i].cell(x, 4).value, tempsheettup[i].cell(x, 5).value, tempsheettup[i].cell(x, 6).value, tempsheettup[i].cell(x, 7).value, tempsheettup[i].cell(x, 8).value, tempsheettup[i].cell(x, 9).value, tempsheettup[i].cell(x, 10).value, tempsheettup[i].cell(x, 11).value)
                            self.student_table.insert('', a, values = val)
                        x+=1
                        a+=1
                    x = 2

                wb.save("C:\\Users\\" + user + "\\Documents\\Student_Database.xlsx")

            except FileNotFoundError:
                messagebox.showerror('No Data Found', 'No Such Data Available.\t\t\n') 
        else:
            messagebox.showerror('No Data Found', 'No Such Data Available.\t\t\n') 



if __name__ == "__main__":
    root = Tk()
    ob = Student(root)
    root.mainloop() 


