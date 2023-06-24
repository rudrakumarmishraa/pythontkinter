#Importing Packages Used
from tkinter import *
from tkinter import ttk
from tkinter import messagebox ,simpledialog
from PIL import Image, ImageTk
import openpyxl as xl
from os import getlogin

user = getlogin()

#########################  Hospital Window  ###########################
class hospital:
    def __init__(self, root):
        self.root = root
        self.root.title("Hospital Management System made by Rudra Kumar Mishra")
        self.root.geometry('1366x786+0+0')
        self.root.minsize(1366, 768)
        self.root.maxsize(1366, 768)
        self.root["bg"]="green"
        lbltitle = Label(self.root, bd = 20, relief=RIDGE, text = "HOSPITAL MANAGEMENT SYSTEM", fg = "white", bg = "black", font = ("times new roman", 50,"bold"))
        lbltitle.pack(side = TOP, fill = X)

                #Variables for Storing Data

        self.Name_of_Tablet=StringVar()
        self.Reference_Number=StringVar()
        self.Dose=StringVar()
        self.Expiry_Date=StringVar()
        self.Daily_Dose=StringVar()
        self.Side_Effects=StringVar()
        self.Patient_Name=StringVar()
        self.Address=StringVar()
        self.Mobile_Number=StringVar()
        self.Date_of_Birth=StringVar()
        self.Problem=StringVar()
        self.Date_of_Checkup=StringVar()
                

                # ----------------Data Frame---------------------------------------
                # Used to collect information of the Patient

        DataFrame = Frame(self.root, border=20, relief=RIDGE)
        DataFrame.place(x=0,y=130, width=1356, height=300)

        DataFrameLeft = LabelFrame(DataFrame, text = "Patient Information", border=10, relief=RIDGE, padx=10, font = ("times new roman", 12,"bold"))
        DataFrameLeft.place(x=0, y=5, width=900, height=250)


        lblNameTablet = Label(DataFrameLeft, text = "Name of tablets", padx=2,pady=6, font = ("bold")).grid(row=0,column=0)

        NameTablet = ttk.Combobox(DataFrameLeft, width=30, textvariable=self.Name_of_Tablet)
        NameTablet["values"]=("Dolo 650", "Vaccine Dose 1", "Vaccine Dose 2", "Abacavir", "Abacavirn", "Abacavir", "Acyclovir", "Alemtuzumab", "Alendronate", "Allopurinol", "Amifostine", "Amikacin", "Aminocaproic Acid", "Amitriptyline", "Amlodipine", "Amoxicillin", "Amoxicillin", "Amphotericin B", "Ampicillin", "Atovaquone", "Azithromycin", "Baclofen", "Bleomycin", "Bortezomib", "Bosentan", "Busulfan", "Calcium", "Captopril", "Carbamazepine", "Carboplatin", "Carmustine", "Cefaclor", "Cefepime", "Cefixime", "Ceftazidime", "Cefuroxime", "Celecoxib", "Cephalexin", "Cidofovir", "Cisplatin", "Cladribine", "Clarithromycin", "Clindamycin", "Clobazam", "Clofarabine", "Codeine", "Crizanlizumab", "Crizotinib", "Cyclobenzaprine", "Cyclophosphamide", "Cyclosporine", "Cyproheptadine", "Cytarabine", "Dacarbazine", "Dactinomycin", "Dapsone", "Darunavir (Prezista®)", "Dasatinib", "Daunorubicin", "Deferasirox (Exjade®)", "Desmopressin (Stimate®)", "Dexamethasone", "Diclofenac", "Didanosine", "Dinutuximab", "Dobutamine", "Dopamine", "Dornase alfa", "Doxorubicin", "Dronabinol", "Efavirenz", "Efavirenz", "Eltrombopag", "Elvitegravir", "Elvitegravir", "Emicizumab", "Emtricitabine (Emtriva®)", "Emtricitabine", "Emtricitabine", "Emtricitabine", "Enalapril", "Enoxaparin", "Erlotinib", "Erythromycin", "Erythropoietin", "Etonogestrel", "Etoposide", "Etravirine", "Factor VIII", "Famciclovir", "Famotidine", "Fidaxomicin", "Fluconazole", "Fludarabine", "Fluorouracil", "Foscarnet", "Furosemide", "G-CSF (Filgrastim)", "Gabapentin", "Ganciclovir", "Gefitinib", "Gemcitabine", "Gemtuzumab ozogamicin", "GM-CSF (Sargramostim)", "Hydralazine", "Hydrocortisone", "Hydromorphone", "Hydroxyurea", "Ifosfamide", "Imatinib", "Imipenem / cilastatin", "Immune globulin", "Irinotecan", "Isotretinoin", "Itraconazole", "Ketoconazole", "L-glutamine", "Labetalol", "Lamivudine", "Levothyroxine", "Linezolid", "Lomustine", "Lorazepam", "Lorlatinib", "Magnesium", "Maraviroc", "Mechlorethamine", "Megestrol acetate", "Meloxicam", "Melphalan", "Meperidine", "Mercaptopurine", "Meropenem", "Mesna", "Methadone", "Methotrexate", "Methylphenidate", "Metronidazole", "Micafungin", "Mitotane", "Mitoxantrone", "Modafinil", "Morphine", "Muromonab – CD3", "Mycophenolate mofetil", "Nelarabine", "Nelfinavir", "Neuromuscular blockers", "Nevirapine", "Norepinephrine", "Omeprazole", "Ondansetron", "Oxycodone", "Paclitaxel", "PEGaspargase", "Pegfilgrastim", "Pemetrexed", "Penicillin VK", "Pentamidine", "Phenobarbital", "Phenytoin", "Phosphorus", "Posaconazole", "Potassium", "Prednisone", "Probenecid", "Procarbazine", "Promethazine", "Promethazine topical gel", "Propoxyphene", "Raltegravir", "Ranitidine", "Rasburicase", "Regorafenib", "Rilpivirine", "Rilpivirine", "Ritonavir", "Rituximab", "Rivaroxaban", "Ruxolitinib", "Sacubitril", "Saquinavir", "Sirolimus", "Sorafenib", "Stavudine", "Sucralfate", "Sugammadex", "Sunitinib", "Tacrolimus", "Temozolomide", "Teniposide", "Tenofovir", "Thioguanine", "Thiotepa", "Tobramycin", "Topotecan", "Tretinoin", "Tretinoin", "Trimethoprim", "Valproic acid", "Vancomycin", "Vinblastine", "Vincristine", "Voriconazole", "Vorinostat", "Voxelotor", "Warfarin", "Zidovudine", "Zidovudine")
        NameTablet.grid(row=0, column=1)

        lblReferenceNo = Label(DataFrameLeft, text = "Reference Number", padx=2,pady=6, font = ("bold")).grid(row=1,column=0)
        ReferenceNoEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=40, textvariable=self.Reference_Number).grid(row=1, column=1)

        lblDose = Label(DataFrameLeft, text = "Dose", padx=2,pady=6, font = ("bold")).grid(row=2,column=0)
        DoseEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=40, textvariable=self.Dose).grid(row=2, column=1)

        lblExpDate = Label(DataFrameLeft, text = "Expiry Date", padx=2,pady=6, font = ("bold")).grid(row=3,column=0)
        ExpDateEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=40, textvariable=self.Expiry_Date).grid(row=3, column=1)

        lblDailyDose = Label(DataFrameLeft, text = "Daily Dose", padx=2,pady=6, font = ("bold")).grid(row=4,column=0)
        DailyDoseEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=40, textvariable=self.Daily_Dose).grid(row=4, column=1)

        lblSideEffects = Label(DataFrameLeft, text = "Side Effects", padx=2,pady=6, font = ("bold")).grid(row=5,column=0)
        SideEffectsEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=40, textvariable=self.Side_Effects).grid(row=5, column=1)

        lblPatient_Name = Label(DataFrameLeft, text = "Patient Name", padx=2,pady=6, font = ("bold")).grid(row=0,column=2)
        Patient_NameNEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=50, textvariable=self.Patient_Name).grid(row=0, column=3)


        lblAddress = Label(DataFrameLeft, text = "Address", padx=2,pady=6, font = ("bold")).grid(row=1,column=2)
        AddressEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=50, textvariable=self.Address).grid(row=1, column=3)

        lblMobile_Number = Label(DataFrameLeft, text = "Mobile Number", padx=2,pady=6, font = ("bold")).grid(row=2,column=2)
        Mobile_NumberEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=50, textvariable=self.Mobile_Number).grid(row=2, column=3)

        lblDate_of_Birth = Label(DataFrameLeft, text = "Date of Birth", padx=2,pady=6, font = ("bold")).grid(row=3,column=2)
        Date_of_BirthEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=50, textvariable=self.Date_of_Birth).grid(row=3, column=3)

        lblProblem = Label(DataFrameLeft, text = "Problem", padx=2,pady=6, font = ("bold")).grid(row=4,column=2)
        ProblemEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=50, textvariable=self.Problem).grid(row=4, column=3)

        lblDate_of_Checkup = Label(DataFrameLeft, text = "Date of Checkup", padx=2,pady=6, font = ("bold")).grid(row=5,column=2)
        Date_of_CheckupEntry = Entry(DataFrameLeft, border=5, relief=SUNKEN, width=50, textvariable=self.Date_of_Checkup).grid(row=5, column=3)

                #Left Data Frame Ends here Collecting Patient Infromation


                #Right Data Frame Code Starts Here Used for Prescription by Doctors

        DataFrameRight = LabelFrame(DataFrame, text = "Search Results", border=10, relief=RIDGE, padx=10, font = ("times new roman", 12,"bold"))
        DataFrameRight.place(x=910, y=5, width=400, height=250)

        self.txtprescription = Text(DataFrameRight, font=("Lucida Handwritting", 15, "italic"), height = 13, border=5, relief=SUNKEN)
        self.txtprescription.pack(side=TOP, anchor="nw", fill = X)
        scroolbar = Scrollbar(DataFrame)
        scroolbar.pack(fill = Y, side = RIGHT)
        scroolbar.config(command=self.txtprescription.yview)
        self.txtprescription.config(yscrollcommand = scroolbar.set)
                #Right Data Frame Code Ends Here Used for Prescription by Doctors



                #----------------------------Button frame------------------------
        ButtonFrame = Frame(self.root, border=20, relief=RIDGE)
        ButtonFrame.place(x=0,y=430, width=1356, height=70)

            
        btn1 = Button(ButtonFrame, text = "Add Prescription", font=("arial", 12, "bold"), width=20, padx=6, bg="green", fg = "white", command=self.AddPrescription).grid(row=0,column=0)
        btn2 = Button(ButtonFrame, text = "Show ALL", font=("arial", 12, "bold"), width=20, padx=6, bg="green", fg = "white", command=self.Show_ALL).grid(row=0,column=1)
        btn3 = Button(ButtonFrame, text = "Clear Info Field", font=("arial", 12, "bold"), width=20, padx=6, bg="green", fg = "white", command=self.Clear_Info).grid(row=0,column=2)
        btn4 = Button(ButtonFrame, text = "Search", font=("arial", 12, "bold"), width=20, padx=6, bg="green", fg = "white", command = self.Khoj).grid(row=0,column=3)
        btn5 = Button(ButtonFrame, text = "Clear Search", font=("arial", 12, "bold"), width=20, padx=6, bg="green", fg = "white", command=self.Clear).grid(row=0,column=4)
        btn6 = Button(ButtonFrame, text = "Quit", font=("arial", 12, "bold"), width=20, padx=6, bg="green", fg = "white", command = self.root.destroy).grid(row=0,column=5)


            #----------------------------Details frame------------------------ 
        self.DetailsFrame = Frame(self.root, border=20, relief=RIDGE)
        self.DetailsFrame.place(x=0,y=500, width=1366, height=200)
        col = ('Name of Tablet', 'Reference Number', 'Dose', 'Expiry Date', 'Daily Dose', 'Side Effects', 'Patient Name',   'Address', 'Mobile Number', 'Date of Birth', 'Problem', 'Date of Checkup')
        scrool = Scrollbar(self.DetailsFrame, orient=VERTICAL)
        self.hospital_table = ttk.Treeview(self.DetailsFrame, height=200, show='headings', columns=col, yscrollcommand = scrool.set)

        self.hospital_table.column('Name of Tablet',width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Reference Number', width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Dose', width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Expiry Date', width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Daily Dose', width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Side Effects', width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Patient Name', width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Address', width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Mobile Number', width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Date of Birth', width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Problem', width = 100, anchor=CENTER, minwidth = 100)
        self.hospital_table.column('Date of Checkup', width = 100, anchor=CENTER, minwidth = 100)

        self.hospital_table.heading('Name of Tablet', text = 'Name of Tablet')
        self.hospital_table.heading('Reference Number', text = 'Reference Number')
        self.hospital_table.heading('Dose', text = 'Dose')
        self.hospital_table.heading('Expiry Date', text = 'Expiry Date')
        self.hospital_table.heading('Daily Dose', text = 'Daily Dose')
        self.hospital_table.heading('Side Effects', text = 'Side Effects')
        self.hospital_table.heading('Patient Name', text = 'Patient Name')
        self.hospital_table.heading('Address', text = 'Address')
        self.hospital_table.heading('Mobile Number', text = 'Mobile Number')
        self.hospital_table.heading('Date of Birth', text = 'Date of Birth')
        self.hospital_table.heading('Problem', text = 'Problem')
        self.hospital_table.heading('Date of Checkup', text = 'Date of Checkup')

        scrool.pack(side = RIGHT, fill=Y)
        self.hospital_table.pack(side=TOP, fill=BOTH)
        scrool.config(command=self.hospital_table.yview)

    def Clear_Info(self):
        #Funciton Clear all ENTRY Fields
        self.Name_of_Tablet.set("")
        self.Reference_Number.set("")
        self.Dose.set("")
        self.Expiry_Date.set("")
        self.Daily_Dose.set("")
        self.Side_Effects.set("")
        self.Patient_Name.set("")
        self.Address.set("")
        self.Mobile_Number.set("")
        self.Date_of_Birth.set("")
        self.Problem.set("")
        self.Date_of_Checkup.set("")

    def Clear(self):
        # Clear Data from Search
        self.txtprescription.delete(1.0, END)

    # This Function is used to add Detail to the spreadsheet
    def AddPrescription(self):
        #Condition if Neccessary fields are empty
        if self.Reference_Number.get() == "" or self.Patient_Name.get() == "" or self.Name_of_Tablet.get() == "":
            messagebox.showerror("Error", "You need to fill complete details!\t\n", parent = self.root)
        else:
            #Trying to open file and add data into it
            try:
                wb = xl.load_workbook("C:\\Users\\" + user + "\\Documents\\Hospital_Database.xlsx")
                sheet = wb["Data"]
                i = 2
                while sheet.cell(i, 1).value != None:
                    i+=1
                sheet.cell(i, 1).value = self.Name_of_Tablet.get()
                sheet.cell(i, 2).value = self.Reference_Number.get()
                sheet.cell(i, 3).value = self.Dose.get()
                sheet.cell(i, 4).value = self.Expiry_Date.get()
                sheet.cell(i, 5).value = self.Daily_Dose.get()
                sheet.cell(i, 6).value = self.Side_Effects.get()
                sheet.cell(i, 7).value = self.Patient_Name.get()
                sheet.cell(i, 8).value = self.Address.get()
                sheet.cell(i, 9).value = self.Mobile_Number.get()
                sheet.cell(i, 10).value = self.Date_of_Birth.get()
                sheet.cell(i, 11).value = self.Problem.get()
                sheet.cell(i, 12).value = self.Date_of_Checkup.get()
                wb.save("C:\\Users\\" + user + "\\Documents\\Hospital_Database.xlsx")

            #If File not found Creating a file to Add data
            except FileNotFoundError:
                #Creating File
                wb = xl.Workbook("C:\\Users\\" + user + "\\Documents\\Hospital_Database.xlsx")
                sheet = wb.create_sheet("Data")
                wb.save("C:\\Users\\" + user + "\\Documents\\Hospital_Database.xlsx")
                wb = xl.load_workbook("C:\\Users\\" + user + "\\Documents\\Hospital_Database.xlsx")
                sheet = wb["Data"]
                i = 2

                sheet.cell(1, 1).value =  "Name of Tablet"
                sheet.cell(1, 2).value =  "Reference Number"
                sheet.cell(1, 3).value =  "Dose"
                sheet.cell(1, 4).value =  "Expiry Date"
                sheet.cell(1, 5).value =  "Daily Dose"
                sheet.cell(1, 6).value =  "Side Effects"
                sheet.cell(1, 7).value =  "Patient Name"
                sheet.cell(1, 8).value =  "Address"
                sheet.cell(1, 9).value =  "Mobile Number"
                sheet.cell(1, 10).value = "Date of Birth"
                sheet.cell(1, 11).value = "Problem"
                sheet.cell(1, 12).value = "Date of Checkup"

                while sheet.cell(i, 1).value != None:
                    i+=1

                #Adding Data to file
                sheet.cell(i, 1).value = self.Name_of_Tablet.get()
                sheet.cell(i, 2).value = self.Reference_Number.get()
                sheet.cell(i, 3).value = self.Dose.get()
                sheet.cell(i, 4).value = self.Expiry_Date.get()
                sheet.cell(i, 5).value = self.Daily_Dose.get()
                sheet.cell(i, 6).value = self.Side_Effects.get()
                sheet.cell(i, 7).value = self.Patient_Name.get()
                sheet.cell(i, 8).value = self.Address.get()
                sheet.cell(i, 9).value = self.Mobile_Number.get()
                sheet.cell(i, 10).value = self.Date_of_Birth.get()
                sheet.cell(i, 11).value = self.Problem.get()
                sheet.cell(i, 12).value = self.Date_of_Checkup.get()
                wb.save("C:\\Users\\" + user + "\\Documents\\Hospital_Database.xlsx")
        self.Show_ALL()

    def Show_ALL(self):
    # Fucntion Creates Tree View table for the Data stored in Exel File
        try:
            wb = xl.load_workbook("C:\\Users\\" + user + "\\Documents\\Hospital_Database.xlsx")
            sheet = wb["Data"]
            i = 2
            j=0
            self.hospital_table.delete(*self.hospital_table.get_children())
            while sheet.cell(i, 1).value != None:
                oadd = (str(sheet.cell(i, 1).value), str(sheet.cell(i, 2).value), str(sheet.cell(i, 3).value), str(sheet.cell(i, 4).    value), str(sheet.cell(i, 5).value), str(sheet.cell(i, 6).value), str(sheet.cell(i, 7).value), str(sheet.cell(i, 8).    value), str(sheet.cell(i, 9).value), str(sheet.cell(i, 10).value), str(sheet.cell(i, 11).value), str(sheet.cell(i,  12).value))
                self.hospital_table.insert('', j, values=oadd)
                i+=1
                j+=1
        except FileNotFoundError:
            messagebox.showwarning("Error", "There is no database to read Add some data.", parent = self.root)
    
    def Khoj(self):
        try:
            a = simpledialog.askstring("Reference Number", "Give the reference number to search\t\t\n", parent = self.root)
            row = 1
            col = 13
            tup = ("Name of Tablet", "Reference Number", "Dose", "Expiry Date", "Daily Dose", "Side Effects", "Patient Name",   "Address", "Mobile Number", "Date of Birth", "Problem", "Date of Checkup")
            wb = xl.load_workbook("C:\\Users\\" + user + "\\Documents\\Hospital_Database.xlsx")
            sheet = wb["Data"]
            while sheet.cell(row, 1).value != None:
                if sheet.cell(row, 2).value == a:
                    for j in range(1, col):
                        self.txtprescription.insert(END, f"{tup[j-1]} is {sheet.cell(row ,j).value}\n")
                    return None
                else:
                    row+=1
            messagebox.showerror("Error", "Data not found", parent = self.root)
        except FileNotFoundError:
            messagebox.showerror("Error", "No DataBase Found Add Some Data", parent = self.root)
root = Tk()

ob = hospital(root)

root.mainloop()    


